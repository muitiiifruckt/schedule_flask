from request import main as get_data_from_link
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl import load_workbook

def get_schedule(data, group_name):
    """
    Извлекает расписание для заданной группы из файла Excel.

    :param file_path: Путь к файлу Excel.
    :param group_name: Название группы (например, "09-141").
    :return: Расписание на неделю в виде DataFrame или сообщение об ошибке.
    """
    try:
        # Загрузка файла

        sheet_name = "Расписание"  # Предположительно основной лист
        df = pd.read_excel(data, sheet_name=sheet_name)
        
        # Поиск группы
        for row_idx, row in df.iterrows():
            for col_idx, cell in enumerate(row):
                if pd.notna(cell) and group_name in str(cell):
                    return row_idx, col_idx

        return f"Группа '{group_name}' не найдена в файле."

    except Exception as e:
        return f"Ошибка обработки файла: {e}"
def split_merged_cells(file_data):
    # Загружаем рабочую книгу
    workbook = load_workbook(file_data)
    

    return workbook
def get_column_from_cell(workbook, start_row, col_index):
    # Загружаем рабочую книгу
    workbook = load_workbook(workbook)
    
    sheet = workbook.active  # Рабочий лист
    for row in sheet.iter_rows(values_only=True):
        if None not in row:
            print(row)
    # Список для хранения данных
    column_data = []

    # Преобразуем номер колонки в букву, если нужно
    for row in range(start_row, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=col_index).value
        column_data.append(cell_value)
    
    return column_data
def main(group):
    data = get_data_from_link()
    print(group)
    row, column = get_schedule(data,group )
    #workbook = split_merged_cells(data)
    result = get_column_from_cell(data, row, column)
    return result
if __name__ == ("__main__"):
    data = main("09-141")    
    